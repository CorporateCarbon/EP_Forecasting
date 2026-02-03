from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import Any, Dict, List, Optional
from helpers.clean_mi_export import clean_mi_export

from openpyxl import load_workbook, Workbook


@dataclass
class AppConfig:
    forecast_workbook: str
    master_inventory_workbook: str
    declared_project_portfolio_workbook: str
    save_master_inventory_output: str
    save_forecast_delta_output: str


REQUIRED_INVENTORY_HEADERS: List[str] = [
    "Name", "Subitems", "Registry ID", "Inventory ID", "Status", "Issuance Status",
    "Days Since (Status)", "Date - Total Amount", "Total Amount (ACCUs)",
    "Realised Amount (ACCUs to CCG)", "Date - Realised Amount", "Forecasted Submission Date",
    "Unit Type", "Application ID", "Reporting Period - Start", "Reporting Period - End",
    "RP", "Delay Flag", "Data Source", "Data Update Date", "Declared Projects Portfolio",
    "Project Number", "Entity", "Proponents", "Methodology", "Business Unit", "Project Stage",
    "Operational Model", "Fee Model", "Number", "Unit", "Item ID", "Key",
]

PORTFOLIO_FIELDS: List[str] = [
    "Name",
    "Subitems",
    "Registry ID",
    "Project ID",          # written to inventory as Project Number
    "Methodology",
    "Project Stage",
    "Proponents",
    "Business Unit",
    "Operational Model",
    "Fee Model",
    "Entity",
    "Number",
    "Unit",
]

FORECAST_TO_INVENTORY_MAP: Dict[str, str] = {
    "RP": "RP",
    "Reporting Period - Start": "Reporting Period - Start",
    "Reporting Period - End": "Reporting Period - End",
    "ACCUs Realised": "Total Amount (ACCUs)",
}


def _norm(x: Any) -> str:
    return "" if x is None else str(x).strip()


def _lower_norm(x: Any) -> str:
    return _norm(x).lower()


def _build_header_map(ws, header_row: int = 1) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        key = _lower_norm(v)
        if key and key not in mapping:
            mapping[key] = col
    return mapping


def _ensure_headers(ws, required_headers: List[str]) -> Dict[str, int]:
    header_map = _build_header_map(ws, 1)
    if not header_map:
        for i, h in enumerate(required_headers, start=1):
            ws.cell(row=1, column=i).value = h
        return _build_header_map(ws, 1)

    next_col = ws.max_column + 1
    for h in required_headers:
        key = _lower_norm(h)
        if key not in header_map:
            ws.cell(row=1, column=next_col).value = h
            header_map[key] = next_col
            next_col += 1
    return header_map


def _find_sheet_by_keywords(wb, keywords: List[str]):
    for name in wb.sheetnames:
        lname = name.lower()
        if all(k.lower() in lname for k in keywords):
            return wb[name]
    return None


def _find_row_by_value(ws, header_map: Dict[str, int], header_name: str, value: Any) -> Optional[int]:
    key = _lower_norm(header_name)
    if key not in header_map:
        raise ValueError(f"Could not find column '{header_name}' in sheet '{ws.title}'.")
    target = _norm(value)
    if not target:
        return None
    col = header_map[key]
    for r in range(2, ws.max_row + 1):
        if _norm(ws.cell(row=r, column=col).value) == target:
            return r
    return None


def _to_datetime(v):
    """
    ALWAYS returns a date or None.
    Handles Excel dates, datetimes, ISO strings, and Excel-serial numbers.
    """
    if v is None:
        return None

    if isinstance(v, datetime):
        return v.date()

    if isinstance(v, date):
        return v

    # Excel serial date
    if isinstance(v, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=int(v))).date()
        except Exception:
            return None

    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass

    return None


def _write(ws, headers: Dict[str, int], row: int, header_name: str, value: Any) -> None:
    key = _lower_norm(header_name)
    if key not in headers:
        return
    ws.cell(row=row, column=headers[key]).value = value


def _row_as_list(ws, row: int, max_col: int) -> List[Any]:
    return [ws.cell(row=row, column=c).value for c in range(1, max_col + 1)]


def _append_table(ws_out, headers: List[str], rows: List[List[Any]]) -> None:
    for c, h in enumerate(headers, start=1):
        ws_out.cell(row=1, column=c).value = h
    for r_idx, row_vals in enumerate(rows, start=2):
        for c, v in enumerate(row_vals, start=1):
            ws_out.cell(row=r_idx, column=c).value = v

def _to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except Exception:
        return 0.0

def add_forecast_to_inventory(config: AppConfig) -> None:
    clean_mi_export(config.master_inventory_workbook)
    clean_mi_export(config.declared_project_portfolio_workbook)    
    run_date = datetime.now().date()  # date only

    # ---- Forecast workbook ----
    f_wb = load_workbook(config.forecast_workbook, data_only=True)
    f_ws = f_wb[f_wb.sheetnames[0]]

    erf = _norm(f_ws["B2"].value)
    if not erf:
        raise ValueError("ERF value was blank. Expected it in cell B2 of the forecast workbook.")

    f_headers = _build_header_map(f_ws, 1)
    for needed in FORECAST_TO_INVENTORY_MAP.keys():
        if _lower_norm(needed) not in f_headers:
            raise ValueError(f"Forecast sheet missing required header '{needed}' in row 1.")

    # Determine cutoff = first forecast RP Start (minimum RP Start among forecast rows)
    rp_start_col = f_headers[_lower_norm("Reporting Period - Start")]
    rp_num_col = f_headers[_lower_norm("RP")]

    forecast_rows: List[Dict[str, Any]] = []
    cutoff_start_dt: Optional[date] = None  # NOTE: this is a date, not datetime

    for r in range(2, f_ws.max_row + 1):
        rp_val = f_ws.cell(row=r, column=rp_num_col).value
        if _norm(rp_val) == "":
            continue

        start_val = f_ws.cell(row=r, column=rp_start_col).value
        start_dt = _to_datetime(start_val)  # returns date
        if start_dt is None:
            raise ValueError(f"Row {r}: Reporting Period - Start missing or not a date.")

        if cutoff_start_dt is None or start_dt < cutoff_start_dt:
            cutoff_start_dt = start_dt

        forecast_rows.append({
            "row_index": r,
            "RP": f_ws.cell(row=r, column=f_headers[_lower_norm("RP")]).value,
            "Reporting Period - Start": start_dt,  # date
            "Reporting Period - End": _to_datetime(f_ws.cell(row=r, column=f_headers[_lower_norm("Reporting Period - End")]).value),  # date
            "ACCUs Realised": f_ws.cell(row=r, column=f_headers[_lower_norm("ACCUs Realised")]).value,
        })

    if cutoff_start_dt is None:
        raise ValueError("No forecast data rows found (RP column was empty).")

    # ---- Declared Projects Portfolio workbook ----
    p_wb = load_workbook(config.declared_project_portfolio_workbook, data_only=True)
    p_ws = _find_sheet_by_keywords(p_wb, ["declared", "portfolio"]) or p_wb[p_wb.sheetnames[0]]
    p_headers = _build_header_map(p_ws, 1)

    if "registry id" not in p_headers:
        raise ValueError(f"Declared Projects Portfolio sheet '{p_ws.title}' missing 'Registry ID' column.")

    p_row = _find_row_by_value(p_ws, p_headers, "Registry ID", erf)
    if p_row is None:
        raise ValueError(f"Could not find ERF '{erf}' in Declared Projects Portfolio 'Registry ID'.")

    # ---- Master inventory workbook ----
    m_wb = load_workbook(config.master_inventory_workbook)
    inventory_ws = _find_sheet_by_keywords(m_wb, ["inventory"]) or m_wb[m_wb.sheetnames[0]]
    i_headers = _ensure_headers(inventory_ws, REQUIRED_INVENTORY_HEADERS)

    # REQUIRED columns for deletion logic
    if _lower_norm("Registry ID") not in i_headers:
        raise ValueError("Inventory sheet missing 'Registry ID' column.")
    if _lower_norm("Reporting Period - End") not in i_headers:
        raise ValueError("Inventory sheet missing 'Reporting Period - End' column.")

    inv_registry_col = i_headers[_lower_norm("Registry ID")]
    inv_rpend_col = i_headers[_lower_norm("Reporting Period - End")]

    # --- Snapshot: all inventory rows where ERF matches Registry ID (BEFORE deletion) ---
    inventory_snapshot_headers = [
        inventory_ws.cell(row=1, column=c).value for c in range(1, inventory_ws.max_column + 1)
    ]
    inv_match_rows_before: List[List[Any]] = []
    for r in range(2, inventory_ws.max_row + 1):
        if _norm(inventory_ws.cell(row=r, column=inv_registry_col).value) == erf:
            inv_match_rows_before.append(_row_as_list(inventory_ws, r, inventory_ws.max_column))

    # --- Delete rows: Registry ID == ERF AND Reporting Period - End > cutoff_start_dt ---
    deleted_rows: List[List[Any]] = []
    kept_rows: List[List[Any]] = []


    for r in range(inventory_ws.max_row, 1, -1):
        if _norm(inventory_ws.cell(row=r, column=inv_registry_col).value) != erf:
            continue

        rp_end_dt = _to_datetime(inventory_ws.cell(row=r, column=inv_rpend_col).value)  # date
        # If RP End is blank/unparseable, treat as NOT deletable (kept)
        if rp_end_dt is None:
            kept_rows.append(_row_as_list(inventory_ws, r, inventory_ws.max_column))
            continue

        if rp_end_dt > cutoff_start_dt:
            deleted_rows.append(_row_as_list(inventory_ws, r, inventory_ws.max_column))
            inventory_ws.delete_rows(r, 1)
        else:
            kept_rows.append(_row_as_list(inventory_ws, r, inventory_ws.max_column))

    kept_rows.reverse()
    deleted_rows.reverse()

    # --- Append new inventory rows for every forecast row ---
    rows_written = 0
    new_forecast_rows: List[List[Any]] = []

    for fr in forecast_rows:
        if fr["Reporting Period - End"] is None:
            raise ValueError(f"Row {fr['row_index']}: Reporting Period - End missing or not a date.")

        out_row = inventory_ws.max_row + 1
        row_dict: Dict[str, Any] = {}


        # portfolio -> inventory
        for field in PORTFOLIO_FIELDS:
            src_key = _lower_norm(field)
            if src_key not in p_headers:
                raise ValueError(f"Declared Projects Portfolio missing required column '{field}'.")

            val = p_ws.cell(row=p_row, column=p_headers[src_key]).value
            dest_field = "Project Number" if src_key == "project id" else field
            _write(inventory_ws, i_headers, out_row, dest_field, val)

        # forecast -> inventory (NOTE: fr values are already DATE objects)
        _write(inventory_ws, i_headers, out_row, "RP", fr["RP"])
        
        _write(
            inventory_ws,
            i_headers,
            out_row,
            "Reporting Period - Start",
            fr["Reporting Period - Start"] + timedelta(days=1),
        )


        _write(inventory_ws, i_headers, out_row, "Reporting Period - End", fr["Reporting Period - End"])
        _write(inventory_ws, i_headers, out_row, "Total Amount (ACCUs)", fr["ACCUs Realised"])

        # derived (date math stays date-only)
        rp_end_dt = fr["Reporting Period - End"]  # date
        _write(inventory_ws, i_headers, out_row, "Forecasted Submission Date", rp_end_dt + timedelta(days=2))
        _write(inventory_ws, i_headers, out_row, "Date - Total Amount", rp_end_dt + timedelta(days=92))

        # fixed
        _write(inventory_ws, i_headers, out_row, "Status", "Forecasted")
        _write(inventory_ws, i_headers, out_row, "Data Update Date", run_date)

        # Capture the full newly-written row in the REQUIRED_INVENTORY_HEADERS order
        new_forecast_rows.append([
            inventory_ws.cell(row=out_row, column=i_headers[_lower_norm(h)]).value if _lower_norm(h) in i_headers else None
            for h in REQUIRED_INVENTORY_HEADERS
        ])

        rows_written += 1

    # --- Lifetime ACCU delta (new total - removed total) ---
    accu_header = "Total Amount (ACCUs)"

    # NEW rows (REQUIRED_INVENTORY_HEADERS order)
    new_idx = {h: i for i, h in enumerate(REQUIRED_INVENTORY_HEADERS)}
    new_total = 0.0
    if accu_header in new_idx:
        col_i = new_idx[accu_header]
        new_total = sum(_to_float(r[col_i]) for r in new_forecast_rows)

    # REMOVED rows (inventory_snapshot_headers order)
    snap_idx = {h: i for i, h in enumerate(inventory_snapshot_headers)}
    removed_total = 0.0
    if accu_header in snap_idx:
        col_i = snap_idx[accu_header]
        removed_total = sum(_to_float(r[col_i]) for r in deleted_rows)

    lifetime_accu_delta = new_total - removed_total

    # --- Save master inventory output ---
    out_master = str(Path(config.save_master_inventory_output))
    Path(out_master).parent.mkdir(parents=True, exist_ok=True)
    m_wb.save(out_master)

    # --- Forecast delta workbook with 3 sheets ---
    delta_wb = Workbook()

    # Sheet 1: Summary
    s1 = delta_wb.active
    s1.title = "Summary"
    s1["A1"] = "Run Date"
    s1["B1"] = "ERF (Registry ID)"
    s1["C1"] = "Cutoff (First RP Start)"
    s1["D1"] = "Rows Deleted"
    s1["E1"] = "Rows Added"
    s1["F1"] = "Lifetime ACCUs Delta"
    s1["A2"] = run_date
    s1["B2"] = erf
    s1["C2"] = cutoff_start_dt  # already a date
    s1["D2"] = len(deleted_rows)
    s1["E2"] = rows_written
    s1["F2"] = lifetime_accu_delta

    # Sheet 2: All ERF-matching inventory rows (before deletion)
    s2 = delta_wb.create_sheet("Old Inventory Rows")
    _append_table(s2, inventory_snapshot_headers, inv_match_rows_before)



    # Sheet 3: Not deleted + forecasts that replace deleted
    # Sheet 3: Keep vs Forecast Replace
    s3 = delta_wb.create_sheet("Keep vs Forecast Replace")

    # --- Section 1: KEPT inventory rows ---
    s3["A1"] = "KEPT inventory rows (Registry ID == ERF, RP End <= cutoff)"
    current_row = 2

    if kept_rows:
        # headers
        for c, h in enumerate(inventory_snapshot_headers, start=1):
            s3.cell(row=current_row, column=c).value = h
        current_row += 1

        # rows
        for row_vals in kept_rows:
            for c, v in enumerate(row_vals, start=1):
                s3.cell(row=current_row, column=c).value = v
            current_row += 1

    # leave a blank gap
    current_row += 2

    # --- Section 2: REMOVED inventory rows ---
    s3.cell(row=current_row, column=1).value = "REMOVED inventory rows (Registry ID == ERF, RP End > cutoff)"
    current_row += 1

    if deleted_rows:
        # headers
        for c, h in enumerate(inventory_snapshot_headers, start=1):
            s3.cell(row=current_row, column=c).value = h
        current_row += 1

        # rows
        for row_vals in deleted_rows:
            for c, v in enumerate(row_vals, start=1):
                s3.cell(row=current_row, column=c).value = v
            current_row += 1

    # leave a blank gap
    current_row += 2

    # --- Section 3: NEW inventory rows to be written (same as New Forecasts) ---
    s3.cell(row=current_row, column=1).value = "NEW inventory rows to be written (replacement set)"
    current_row += 1

    # headers (inventory-shaped)
    for c, h in enumerate(REQUIRED_INVENTORY_HEADERS, start=1):
        s3.cell(row=current_row, column=c).value = h
    current_row += 1

    # rows
    for row_vals in new_forecast_rows:
        for c, v in enumerate(row_vals, start=1):
            s3.cell(row=current_row, column=c).value = v
        current_row += 1





    s4 = delta_wb.create_sheet("New Inventory Rows")

    # Headers
    for c, h in enumerate(REQUIRED_INVENTORY_HEADERS, start=1):
        s4.cell(row=1, column=c).value = h

    current_row = 2

    # --- KEPT inventory rows first ---
    if kept_rows:
        # Build a mapping from snapshot headers -> index
        snap_index = {h: i for i, h in enumerate(inventory_snapshot_headers)}

        for row_vals in kept_rows:
            for c, h in enumerate(REQUIRED_INVENTORY_HEADERS, start=1):
                if h in snap_index:
                    s4.cell(row=current_row, column=c).value = row_vals[snap_index[h]]
            current_row += 1

    # --- NEW inventory rows next ---
    for row_vals in new_forecast_rows:
        for c, v in enumerate(row_vals, start=1):
            s4.cell(row=current_row, column=c).value = v
        current_row += 1

    # Save delta output
    out_delta = str(Path(config.save_forecast_delta_output))
    Path(out_delta).parent.mkdir(parents=True, exist_ok=True)
    delta_wb.save(out_delta)


def run_process(config: AppConfig) -> None:
    add_forecast_to_inventory(config)
