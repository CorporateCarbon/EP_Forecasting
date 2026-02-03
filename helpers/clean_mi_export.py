#%%##
from __future__ import annotations

from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _row_is_empty(ws: Worksheet, row_idx: int) -> bool:
    """Return True if *all* cells in the given 1-based row are empty."""
    max_col = ws.max_column
    if max_col is None or max_col == 0:
        return True
    for cell in ws[row_idx]:
        if cell.value not in (None, ""):
            return False
    return True


def _strip_images_on_rows(ws: Worksheet, max_row_to_strip: int = 2) -> None:
    """Remove images anchored to the first `max_row_to_strip` rows (in-place).

    This is conservative and only used when we actually delete those rows.
    """
    images = getattr(ws, "_images", None)
    if not images:
        return

    keep = []
    for img in images:
        try:
            # openpyxl uses 0-based row indices in anchors
            anchor = getattr(img, "anchor", None)
            if anchor is None or not hasattr(anchor, "_from"):
                keep.append(img)
                continue
            row_from = anchor._from.row + 1  # convert to 1-based
        except Exception:
            keep.append(img)
            continue

        if row_from > max_row_to_strip:
            keep.append(img)

    ws._images = keep


def clean_mi_export(xlsx_path, sheet_name: Optional[str] = None, output_path: Optional[str | Path] = None):
    """Clean a Monday.com Master Inventory export **in-place on the file**.

    Behaviour:
    - Load the workbook with openpyxl.
    - Select the target sheet (explicit name or active sheet).
    - If *both* the first and second rows are completely empty, then:
        * Delete the first 2 rows.
        * Remove any images anchored to those rows.
    - Otherwise:
        * Leave the sheet structure unchanged, but perform a light sanity
          check that the first row looks like a header (all non-empty cells
          are strings). The check is informational only; no exception is raised.

    Parameters
    ----------
    xlsx_path : str or Path
        Input Excel file (Monday export).
    sheet_name : str, optional
        Sheet to clean. Defaults to the active sheet.
    output_path : str or Path, optional
        Where to write the cleaned workbook. If None, overwrite `xlsx_path`.

    Returns
    -------
    Path
        Path to the cleaned workbook.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Input Excel file not found: {xlsx_path}")

    wb = load_workbook(xlsx_path)
    ws: Worksheet
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found in workbook {xlsx_path.name}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Guard for very small / empty sheets
    if ws.max_row is None or ws.max_row == 0:
        out_path = Path(output_path) if output_path else xlsx_path
        wb.save(out_path)
        return out_path

    # Decide whether to remove the first 2 rows

    # Find the header row where column A == "Name"
    header_row_idx = None
    for row_idx in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=1).value
        if isinstance(cell_val, str) and cell_val.strip() == "Name":
            header_row_idx = row_idx
            break

    # Delete everything above the header row
    if header_row_idx and header_row_idx > 1:
        _strip_images_on_rows(ws, max_row_to_strip=header_row_idx - 1)
        ws.delete_rows(1, header_row_idx - 1)

    out_path = Path(output_path) if output_path else xlsx_path
    wb.save(out_path)
    return out_path


def clean_master_inventory_export(df: pd.DataFrame) -> pd.DataFrame:
    """Clean a Master Inventory DataFrame that came from a Monday export.

    Behaviour:
    - If the first two **data rows** are completely NaN/empty, drop them.
    - Otherwise, return the DataFrame unchanged.

    This is intended for use in the pandas-based pipeline (main.py),
    which already uses the first non-empty row as headers via `read_excel`.
    """
    if df is None or df.empty:
        return df

    out = df.copy()

    if len(out) >= 2:
        first_two = out.iloc[:2]
        # Treat NaN/None/empty-string as empty
        is_empty = first_two.replace("", pd.NA).isna().all(axis=1)
        if bool(is_empty.all()):
            out = out.iloc[2:].reset_index(drop=True)

    return out


if __name__ == "__main__":
    # Example manual usage; adjust paths as needed or comment out.
    # clean_mi_export(
    #     r"C:\path\to\Master_Inventory_export.xlsx",
    #     sheet_name=None,
    #     output_path=None,  # Overwrite in-place
    # )
    pass

# %%
