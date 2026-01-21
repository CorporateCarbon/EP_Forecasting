# -*- coding: utf-8 -*-
"""
Schedule 4 Forecaster — clean Equation 16 implementation + issuance discount

What this script produces (per reporting period):
1) CP_i (method carbon stock) using Equation 16 (includes DPP from Eq16)
2) ΔCP_i (method abatement for the RP) = CP_i(t) - CP_i(t-1)
3) Issued ACCUs (optional) = ΔCP_i * ISSUANCE_FACTOR (e.g. 0.75 for 20%+5%)
4) Optional first-year deduction (project-specific) applied to ISSUED ACCUs in RP1

IMPORTANT:
- Equation 16 CP_i is defined in tonnes CO2-e. Ensure your inputs match.
- If your calculator values are already tCO2-e, set INPUT_UNITS="CO2E".
- If your calculator values are tonnes of Carbon (tC), set INPUT_UNITS="C" so it converts to CO2-e.

Optional:
- Can pull CBASE/CLT from an Excel calculator via openpyxl if you provide a path/sheet/cells.
"""
#%%##
from __future__ import annotations
from dataclasses import dataclass
from datetime import date, datetime
from typing import List, Tuple, Optional
import csv
from dateutil.relativedelta import relativedelta

# -----------------------------
# Manual inputs (or read from Excel)
# -----------------------------
# From your Sch4 calculator:
# #FC2024
# CBASE = 596.78        # Calculations!B21  (net baseline carbon stock for CEA, Eq6 input)
# CLT   = 184419.56     # Calculations!B28  (predicted long-term project scenario carbon stock, Eq9 input)
 
#FC2016
CBASE = 418.14 #value from b21
CLT   = 149697.25  #Value from B28
PERMANENCE_YEARS = 25
EQ16_DPP = 0.75 if PERMANENCE_YEARS == 25 else 1.0   # DPP used INSIDE Equation 16

OUT_PATH = r'C:\Users\GeorginaDoyle\github\EP_Forecasting\PF_2020'
# Separate issuance discount for 25-year projects:
# 20% permanence discount + 5% buffer => 25% total => multiply issued units by 0.75
ISSUANCE_FACTOR = 0.75 if PERMANENCE_YEARS == 25 else 1.0

# CEA definition date (used for month counting in Eq16)
CEA_START = date(2021, 6, 25)

# Reporting period schedule (annual cadence here; you can make it more frequent if needed)
FIRST_RP_START = date(2025, 10, 31)
FIRST_RP_END   = date(2026, 6, 30)
LAST_RP_END    = date(2046, 6, 30)

# Project-specific adjustment (apply to ISSUED ACCUs in first RP only)
FIRST_YEAR_DEDUCTION_ISSUED = 12821.0  # ACCUs (tCO2-e). Set 0 to disable.

# Units handling
# - Equation 16 CP_i is in tonnes CO2-e
# - If your CBASE/CLT are already CO2-e (common in calculators), set "CO2E"
# - If they are tonnes C, set "C" to convert to CO2-e
INPUT_UNITS = "CO2E"  # "CO2E" or "C"

# If CBASE/CLT are per hectare and you want totals:
AREA_HA: Optional[float] = None  # e.g., 1234.56; leave None if already totals

# -----------------------------
# Optional: read from Excel calculator
# -----------------------------
def read_from_excel(
    excel_path: str,
    sheet_name: str = "Calculations",
    cbase_cell: str = "B21",
    clt_cell: str = "B28",
) -> Tuple[float, float]:
    """
    Read CBASE and CLT from an Excel file using openpyxl.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("openpyxl is required for Excel reading. pip install openpyxl") from e

    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    cbase = ws[cbase_cell].value
    clt = ws[clt_cell].value
    if cbase is None or clt is None:
        raise ValueError(f"Missing values in Excel: {sheet_name}!{cbase_cell} or {sheet_name}!{clt_cell}")
    return float(cbase), float(clt)

# -----------------------------
# Core logic
# -----------------------------
CO2_PER_C = 44.0 / 12.0

def to_co2e(x: float, input_units: str) -> float:
    if input_units.upper() == "CO2E":
        return float(x)
    if input_units.upper() == "C":
        return float(x) * CO2_PER_C
    raise ValueError(f"Unknown INPUT_UNITS={input_units!r}; use 'CO2E' or 'C'")

def months_completed(d1: date, d2: date) -> int:
    """
    Whole months completed between d1 and d2.
    Matches typical "completed months" logic for Eq16's 'n' (capped at 180).
    """
    if d2 < d1:
        return -months_completed(d2, d1)
    y = d2.year - d1.year
    m = d2.month - d1.month
    total = y * 12 + m
    if d2.day < d1.day:
        total -= 1
    return total

def eq16_cp(
    cbase_co2e: float,
    clt_co2e: float,
    n_months: int,
    eq16_dpp: float,
) -> float:
    """
    Equation 16:
    CP_i = CBASE + (n/180) * (CLT - CBASE) * DPP
    where n = min(months_completed, 180)
    """
    n = max(0, min(int(n_months), 180))
    return cbase_co2e + (n / 180.0) * (clt_co2e - cbase_co2e) * eq16_dpp

def generate_rp_schedule(first_start: date, first_end: date, last_end: date) -> List[Tuple[date, date]]:
    sched: List[Tuple[date, date]] = []
    rp_start, rp_end = first_start, first_end
    while rp_end <= last_end:
        sched.append((rp_start, rp_end))
        rp_start = rp_start + relativedelta(years=1)
        rp_end   = rp_end   + relativedelta(years=1)
    return sched

@dataclass
class RPRow:
    rp_index: int
    rp_start: date
    rp_end: date
    months_since_cea_defined: int
    cp_method_co2e: float             # CP_i from Eq16 (already includes Eq16 DPP)
    delta_cp_method_co2e: float       # ΔCP_i (method abatement for RP)
    issued_accus_co2e: float          # ΔCP_i * issuance factor (and RP1 deduction if applied)

def forecast() -> List[RPRow]:
    # Convert + scale inputs
    cbase = to_co2e(CBASE, INPUT_UNITS)
    clt   = to_co2e(CLT,   INPUT_UNITS)

    if AREA_HA is not None:
        cbase *= AREA_HA
        clt   *= AREA_HA

    schedule = generate_rp_schedule(FIRST_RP_START, FIRST_RP_END, LAST_RP_END)

    rows: List[RPRow] = []
    prev_cp = None

    for i, (rp_start, rp_end) in enumerate(schedule, start=1):
        n = months_completed(CEA_START, rp_end)

        cp = eq16_cp(
            cbase_co2e=cbase,
            clt_co2e=clt,
            n_months=n,
            eq16_dpp=EQ16_DPP,
        )

        delta_cp = cp if prev_cp is None else (cp - prev_cp)

        # Apply issuance discount (20% permanence + 5% buffer) to convert method abatement to "issued ACCUs"
        issued = delta_cp * ISSUANCE_FACTOR

        # Apply project-specific RP1 deduction at the ISSUED level (only first reporting period)
        if i == 1 and FIRST_YEAR_DEDUCTION_ISSUED:
            issued = issued - float(FIRST_YEAR_DEDUCTION_ISSUED)

        rows.append(RPRow(
            rp_index=i,
            rp_start=rp_start,
            rp_end=rp_end,
            months_since_cea_defined=n,
            cp_method_co2e=cp,
            delta_cp_method_co2e=delta_cp,
            issued_accus_co2e=issued,
        ))

        prev_cp = cp

    return rows

def write_csv(path: str, rows: List[RPRow]) -> str:
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "RP",
            "RP_Start",
            "RP_End",
            "Months_since_CEA_defined",
            "CP_Method_CO2e (Eq16, includes Eq16 DPP)",
            "Delta_CP_Method_CO2e",
            "Issued_ACCUs_CO2e (Delta * issuance_factor, minus RP1 deduction if set)",
        ])
        for r in rows:
            w.writerow([
                r.rp_index,
                r.rp_start.isoformat(),
                r.rp_end.isoformat(),
                r.months_since_cea_defined,
                round(r.cp_method_co2e, 6),
                round(r.delta_cp_method_co2e, 6),
                round(r.issued_accus_co2e, 6),
            ])
    return path

def main():
    rows = forecast()
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    out = OUT_PATH + f"sch4_FC16_forecast_{ts}.csv"
    write_csv(out, rows)

    print("\n--- Settings ---")
    print(f"INPUT_UNITS={INPUT_UNITS}, AREA_HA={AREA_HA}")
    print(f"EQ16_DPP={EQ16_DPP} (inside Eq16 CP)")
    print(f"ISSUANCE_FACTOR={ISSUANCE_FACTOR} (20%+5% issuance discount)")
    print(f"FIRST_YEAR_DEDUCTION_ISSUED={FIRST_YEAR_DEDUCTION_ISSUED}\n")

    print(f"Saved: {out}\n")
    print("First 3 rows preview:")
    for r in rows[:3]:
        print(
            f"RP{r.rp_index}: end={r.rp_end} n={r.months_since_cea_defined} "
            f"CP={r.cp_method_co2e:.2f} ΔCP={r.delta_cp_method_co2e:.2f} Issued={r.issued_accus_co2e:.2f}"
        )

if __name__ == "__main__":
    main()

# %%
