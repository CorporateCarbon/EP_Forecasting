from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple, Any, Dict

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.datetime import from_excel
from dateutil.relativedelta import relativedelta

# ---------- Config container (matches your GUI config style) ----------
@dataclass

class EngineConfig:
    starting_rp_number: int
    rp_length_months: int
    start_year: int
    start_month: int
    start_day: int

    forecast_full_lifecycle: bool
    forecast_number_of_rps: int | None

    input_calculator_file: str
    save_aggregated_output: str


# ---------- Engine ----------
class ForecastEngine:
    """
    Loads the calculator workbook and indexes the 'Forecast_script_helper' sheet's Column A labels
    for fast repeated writes/reads across many iterations.
    """

    TARGET_SHEET = "Forecast_script_helper"

    # Labels (Column A) we care about
    LABEL_CURRENT_RP = "Current RP"
    LABEL_RP_END_YEAR = "Current RP End Year"
    LABEL_RP_END_MONTH = "current rp end month"
    LABEL_RP_END_DAY = "current rp end day"
    LABEL_ACCUS_REALISED = "ACCUs Realised"
    LABEL_RP_LENGTH = "RP Length"


    def __init__(self, input_calculator_file: str):
        self.input_path = Path(input_calculator_file)
        if not self.input_path.exists():
            raise FileNotFoundError(f"Input calculator file not found: {self.input_path}")

        # data_only=False so we can write and see formulas if present
        self.wb = load_workbook(self.input_path, data_only=False)
        if self.TARGET_SHEET not in self.wb.sheetnames:
            raise ValueError(
                f"Worksheet '{self.TARGET_SHEET}' not found in {self.input_path.name}. "
                f"Available sheets: {self.wb.sheetnames}"
            )

        self.ws: Worksheet = self.wb[self.TARGET_SHEET]

        # Build an index of normalized label -> row number for fast lookup
        self.label_row: Dict[str, int] = self._index_labels_in_column_a(self.ws)

        # Validate we can find the key labels up front (fail fast)
        for required in (
            self.LABEL_CURRENT_RP,
            self.LABEL_RP_END_YEAR,
            self.LABEL_RP_END_MONTH,
            self.LABEL_RP_END_DAY,
            self.LABEL_ACCUS_REALISED,
            self.LABEL_RP_LENGTH,
        ):
            if self._norm(required) not in self.label_row:
                raise ValueError(
                    f"Could not find label '{required}' in column A of '{self.TARGET_SHEET}'."
                )

    @staticmethod
    def _norm(x: Any) -> str:
        """Normalize labels for matching: string, stripped, lowercased."""
        if x is None:
            return ""
        return str(x).strip().lower()

    @staticmethod
    def _index_labels_in_column_a(ws: Worksheet) -> Dict[str, int]:
        """
        Scan column A once and map normalized cell text -> row number.
        If duplicates exist, the first occurrence is kept.
        """
        mapping: Dict[str, int] = {}
        max_row = ws.max_row or 1

        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=1).value  # column A
            key = ForecastEngine._norm(v)
            if key and key not in mapping:
                mapping[key] = r
        return mapping

    @staticmethod
    def _coerce_stripped_value(v: Any) -> Any:
        """
        Ensure we strip whitespace if v is a string. If numeric, leave numeric.
        User requested: "strip any blank spaces before inputting these values."
        """
        if isinstance(v, str):
            return v.strip()
        return v

    def write_inputs_and_get_accus(
        self,
        starting_rp_number: int,
        start_year: int,
        start_month: int,
        start_day: int,
        rp_length_months: int,
    ) -> Tuple[datetime, Any]:
        """
        Writes input values next to their labels (col B), then returns:
          (datetime(start_year,start_month,start_day), value_in_colB_of_ACCUs_Realised_row)

        Designed to be called many times in a loop:
        - Uses pre-indexed label rows
        - Only touches a few cells per call
        """

        # Build the datetime we will return
        dt = datetime(int(start_year), int(start_month), int(start_day))

        # Lookups are O(1) because of the pre-index
        row_current_rp = self.label_row[self._norm(self.LABEL_CURRENT_RP)]
        row_year = self.label_row[self._norm(self.LABEL_RP_END_YEAR)]
        row_month = self.label_row[self._norm(self.LABEL_RP_END_MONTH)]
        row_day = self.label_row[self._norm(self.LABEL_RP_END_DAY)]
        row_rp_length = self.label_row[self._norm(self.LABEL_RP_LENGTH)]
        # Write values into column B (one column to the right)
        # Strip any whitespace before writing (as requested)
        self.ws.cell(row=row_current_rp, column=2).value = self._coerce_stripped_value(starting_rp_number)
        self.ws.cell(row=row_year, column=2).value = self._coerce_stripped_value(start_year)
        self.ws.cell(row=row_month, column=2).value = self._coerce_stripped_value(start_month)
        self.ws.cell(row=row_day, column=2).value = self._coerce_stripped_value(start_day)
        self.ws.cell(row=row_rp_length, column=2).value = self._coerce_stripped_value(rp_length_months)
        # Now fetch "ACCUs Realised" from column B
        row_accus = self.label_row[self._norm(self.LABEL_ACCUS_REALISED)]
        accus_value = self.ws.cell(row=row_accus, column=2).value

        return dt, accus_value
    
    def get_project_start_date(self) -> datetime:
        ws = self.ws

        for r in range(1, ws.max_row + 1):
            label = ws.cell(row=r, column=4).value
            if label and str(label).strip().lower() == "project start date":
                val = ws.cell(row=r, column=5).value

                if isinstance(val, datetime):
                    return val
                if isinstance(val, (int, float)):
                    return from_excel(val)

                raise ValueError("Project Start Date is not a valid Excel date.")

        raise ValueError("Could not find 'Project Start Date' in column D.")

    def save_calculator_copy(self, output_path: str) -> None:
        """
        Optional helper: if you want to save a working copy of the calculator after updates.
        """
        self.wb.save(output_path)

    def close(self) -> None:
        self.wb.close()


# ---------- Aggregated workbook creation ----------
def create_aggregated_workbook(save_aggregated_output: str) -> None:
    """
    Creates a new workbook and saves it to the specified path.
    """
    out_path = Path(save_aggregated_output)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Aggregated"

    # Minimal headers (optional; you can change later)
    ws["A1"] = "Date"
    ws["B1"] = "ACCUs Realised"

    wb.save(out_path)
    wb.close()


def run_engine(config: EngineConfig) -> None:
    create_aggregated_workbook(config.save_aggregated_output)
    engine = ForecastEngine(config.input_calculator_file)

    try:
        # Decide n_rps
        if config.forecast_number_of_rps is not None:
            n_rps = int(config.forecast_number_of_rps)
        else:
            project_start_date = engine.get_project_start_date()
            project_end_date = project_start_date + relativedelta(years=25)

            current_date = datetime(config.start_year, config.start_month, config.start_day)
            current_to_end_months = (project_end_date.year - current_date.year) * 12 + (project_end_date.month - current_date.month)

            # number of RPs = months / rp_length_months
            n_rps = current_to_end_months // int(config.rp_length_months)
            if current_to_end_months % int(config.rp_length_months) != 0:
                n_rps += 1  # include partial final RP

        # Example call (you’ll loop this later)
        dt, accus = engine.write_inputs_and_get_accus(
            starting_rp_number=config.starting_rp_number,
            rp_length_months=config.rp_length_months,
            start_year=config.start_year,
            start_month=config.start_month,
            start_day=config.start_day,
        )

        print("n_rps:", n_rps)
        print("Returned:", dt, accus)

    finally:
        engine.close()
