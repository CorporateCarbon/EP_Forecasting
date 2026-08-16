"""
Corrected generalised EP2014 forecaster (independent, openpyxl-based).

Crediting model (CONFIRMED against the signed RP9 Part B calculator, KD review 2026-08-15):
  * Carbon stock  = sum_CEA[ area_ha * (C_trees + C_debris) ] * 44/12          (Eq 11b)
  * Crediting is CUMULATIVE against the last *credited* carbon level (the high-water
    baseline = CV, the previous credited offsets report's CP). A reporting period whose
    stock is BELOW that baseline earns ZERO and does NOT lower the baseline - the project
    must climb back above the previously-credited level before it earns again. This is the
    calculator's own behaviour: CV stays anchored at the pre-decline report (e.g. 23,476.67)
    and A = CP(period end) - CV - emissions, so a post-decline recovery is credited only for
    the genuine excess above the prior credited maximum (no double-crediting of regrowth).
  * cumulative_net(t) = (stock(t) - baseline) - cumulative_emissions_since_baseline
  * ACCUs credited to date = max(0, cumulative_net) * 0.75   (25% buffer on the positive excess)
  * ACCUs issued in period t = increment in credited-to-date (never negative in a forecast)
  * RP9 (reported/locked) taken from the calculator; forecast periods use EB=EF=CN=0
  * Crediting-period END enforced; final RP truncated to a PARTIAL period ending on it.

Output rules (KD review): ACCUs Issued is floored at 0; the running deficit (unissued
negative net) is recorded in its own column; the TOTAL sums ISSUED ACCUs only.

INTERNAL forecasting reproduction only - verify against CER before any external use.
"""
import openpyxl
from datetime import date

C_TO_CO2 = 44/12
BUFFER = 0.75
CREDITING_END = date(2036, 6, 30)   # ERF101808 registered crediting-period end (cer.gov.au)

def cea_series(ws):
    s = {}
    for r in range(10, ws.max_row + 1):
        y, m = ws.cell(r,1).value, ws.cell(r,2).value
        tr, de = ws.cell(r,5).value, ws.cell(r,6).value
        if isinstance(y,(int,float)) and isinstance(m,(int,float)) and isinstance(tr,(int,float)):
            s[(int(y),int(m))] = (tr or 0) + (de or 0)
    return s

def load_part(path, cea_area):
    wb = openpyxl.load_workbook(path, data_only=True)
    series = {name: cea_series(wb[name]) for name in cea_area}
    na = wb["Net Abatement"]
    anchors = {r[0].value: r[2].value for r in na.iter_rows(min_row=5,max_row=11,max_col=4) if r[0].value}
    return series, anchors

def stock(series, cea_area, y, m):
    tot = 0.0
    for name, area in cea_area.items():
        s = series[name]; v = s.get((y,m))
        if v is None:
            keys=[k for k in s if (k[0],k[1])<=(y,m)]
            v = s[max(keys)] if keys else s[max(s)]
        tot += area * v * C_TO_CO2
    return tot

def forecast_part(path, cea_area, reported_rp9, first_forecast_year=2026, first_forecast_rp=10):
    """
    Returns list of dicts per RP with: rp, year, net_period (annual net abatement),
    cum_position (cumulative net vs baseline; the running deficit while negative),
    issued (ACCUs), note.
    """
    series, anchors = load_part(path, cea_area)
    baseline = anchors["CV"]                     # last credited report's CP (high-water)
    emissions_rp9 = anchors["EB"] + anchors["EF"]  # reported RP9 emissions (real)
    cum_emissions = emissions_rp9
    credited = 0                                 # ACCUs credited to date
    rows = []
    # RP9 (reported / locked)
    cp9 = anchors["CP"]
    cum9 = (cp9 - baseline) - cum_emissions
    rows.append(dict(rp="9", year=2025, net_period=reported_rp9["A"], cum=cum9,
                     issued=reported_rp9["ACCUs"], note="reported / locked"))
    credited = max(0, reported_rp9["ACCUs"])
    prev_stock = cp9
    rp = first_forecast_rp
    for y in range(first_forecast_year, CREDITING_END.year+1):
        if y < CREDITING_END.year:
            ey, em, label = y, 12, "full year"
        else:
            ey, em, label = CREDITING_END.year, CREDITING_END.month, "PARTIAL -> 30 Jun 2036"
        s_end = stock(series, cea_area, ey, em)
        net_period = s_end - prev_stock                    # annual net (EB=EF=CN=0 in forecast)
        cum = (s_end - baseline) - cum_emissions           # cumulative vs high-water baseline
        credited_now = round(max(0.0, cum) * BUFFER)
        issued = credited_now - credited
        if issued < 0: issued = 0                          # never claw back in a forecast
        credited = credited_now
        rows.append(dict(rp=str(rp), year=ey, net_period=net_period, cum=cum,
                         issued=issued, note=label))
        prev_stock = s_end
        rp += 1
        if (ey,em) >= (CREDITING_END.year, CREDITING_END.month):
            break
    return rows

PART_A_AREA = {"JIN_CEA_01":6.677401,"JIN_CEA_01A":2.227761,"HID_CEA_01":2.698047,"LIS_CEA_01":0.641232,"LIS_CEA_01A":4.512582}
PART_B_AREA = {"STE_CEA_01":9.2866,"STE_CEA_01A":0.8901,"STE_CEA_02":54.3194,"STE_CEA_02A":7.9903,"STE_CEA_03":12.5641,"STE_CEA_03A":6.9491}

if __name__ == "__main__":
    pathA="/mnt/user-data/uploads/3. FY27 Forecasts/260806_Dogwood Project Calc RP9_Part A.xlsx"
    pathB="/mnt/user-data/uploads/2026 Offsets Report (RP09 - 250101 - 251231)/Report Part_B/Application Docs/02. Abatement Calculator/02. Dogwood Project Calc RP9_Part B.xlsx"
    for tag,path,area,rep in [("PART A",pathA,PART_A_AREA,{"A":427.389,"ACCUs":321}),
                              ("PART B",pathB,PART_B_AREA,{"A":-1672.16,"ACCUs":0})]:
        rows=forecast_part(path,area,rep)
        print("\n"+"="*84+f"\n{tag}")
        print(f"{'RP':>4}{'Yr':>6}{'NetPeriod':>12}{'CumVsBase':>12}{'Issued':>8}  note")
        tot=0
        for r in rows:
            tot+=r['issued']
            print(f"{r['rp']:>4}{r['year']:>6}{r['net_period']:>12.1f}{r['cum']:>12.1f}{r['issued']:>8}  {r['note']}")
        print(f"  TOTAL ISSUED = {tot}")
